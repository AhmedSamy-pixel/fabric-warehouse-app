"""
Fabric Warehouse Inventory Tracking System
===========================================
A single-file Streamlit application with SQLite persistence and Gemini-powered
AI OCR for automatic fabric roll tag reading.

SETUP
-----
1. Install dependencies:
   pip install streamlit google-genai openpyxl pillow

2. Create a `.streamlit/secrets.toml` file next to this app with:
   GEMINI_API_KEY = "your-gemini-api-key-here"

3. Run:
   streamlit run app.py

Author: Senior Python & Streamlit Developer
"""

import io
import os
import re
import json
import uuid
import sqlite3
from datetime import datetime

import streamlit as st
from PIL import Image

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------
DB_PATH = "warehouse_system.db"
IMAGE_DIR = "roll_images"
DEFAULT_SUPPLIER = "El-basha"
GEMINI_MODEL = "gemini-2.5-flash"

# ---------------------------------------------------------------------------
# PAGE CONFIG & GLOBAL CSS
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Fabric Warehouse Inventory",
    page_icon="🧵",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
        /* Force-fix mirrored rear camera preview on mobile devices */
        video {
            transform: scaleX(1) !important;
            -webkit-transform: scaleX(1) !important;
        }
        /* Tidy up camera / uploader widget spacing */
        [data-testid="stCameraInput"] video {
            border-radius: 10px;
        }
        .stButton > button {
            border-radius: 8px;
            font-weight: 600;
        }
        [data-testid="stMetricValue"] {
            font-size: 1.4rem;
        }
        div[data-testid="stForm"] {
            border: 1px solid rgba(49, 51, 63, 0.15);
            border-radius: 10px;
            padding: 1.2rem;
        }
        .roll-id-badge {
            display: inline-block;
            background-color: #2F5233;
            color: white;
            padding: 4px 12px;
            border-radius: 6px;
            font-weight: 600;
            font-size: 0.9rem;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# DATABASE LAYER
# ---------------------------------------------------------------------------
def get_connection() -> sqlite3.Connection:
    """Return a new SQLite connection with row access by column name."""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    """Create tables (if missing) and seed the default supplier."""
    os.makedirs(IMAGE_DIR, exist_ok=True)
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_name TEXT UNIQUE NOT NULL,
                description TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS inventory (
                roll_id TEXT PRIMARY KEY,
                supplier_name TEXT NOT NULL,
                fabric_name TEXT,
                color_shade TEXT,
                pc_no TEXT,
                lot_no TEXT,
                metres REAL,
                weight_kg REAL,
                image_path TEXT,
                status TEXT DEFAULT 'IN_STOCK',
                date_added TIMESTAMP
            )
            """
        )
        cur.execute(
            "SELECT COUNT(*) FROM suppliers WHERE supplier_name = ?",
            (DEFAULT_SUPPLIER,),
        )
        if cur.fetchone()[0] == 0:
            cur.execute(
                "INSERT INTO suppliers (supplier_name, description) VALUES (?, ?)",
                (DEFAULT_SUPPLIER, "Default supplier"),
            )
        conn.commit()
    except Exception as exc:  # pragma: no cover - defensive
        st.error(f"Database initialization failed: {exc}")
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def clean_numeric_string(value) -> str:
    """Extract the first numeric token from a messy OCR string, e.g. '12.5 m' -> '12.5'."""
    if value is None:
        return "0"
    text = str(value).strip()
    if not text:
        return "0"
    match = re.search(r"[-+]?\d*\.?\d+", text)
    return match.group(0) if match else "0"


def safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_jpeg_bytes(raw_bytes: bytes, quality: int = 90) -> bytes:
    """Normalize any image bytes to clean RGB JPEG bytes (avoids RGBA / encoding issues)."""
    img = Image.open(io.BytesIO(raw_bytes))
    if img.mode != "RGB":
        img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality)
    return buf.getvalue()


def generate_roll_id() -> str:
    """Generate a collision-resistant, human-readable roll ID."""
    return f"R{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4].upper()}"


# ---------------------------------------------------------------------------
# GEMINI OCR ENGINE
# ---------------------------------------------------------------------------
OCR_PROMPT = """You are an expert OCR system reading a fabric roll identification tag \
in a textile warehouse. Carefully read all visible text, handwriting, and printed \
labels on the tag in the image.

Extract exactly these six fields:
- fabric: the fabric name / type
- shade: the color or shade name/code
- pc_no: the piece number (PC No.)
- lot_no: the lot number
- metres: the length in metres (numbers only, no units)
- weight: the weight in kilograms (numbers only, no units)

Rules:
- If a field is not visible or unreadable, return an empty string "" for it.
- For metres and weight, return ONLY the numeric value as a string (e.g. "45.5"), no units or extra text.
- Do not guess values that are not present on the tag.
- Respond with STRICT raw JSON only, no markdown code fences, no explanations, no extra text.

Return exactly this JSON shape:
{"fabric": "", "shade": "", "pc_no": "", "lot_no": "", "metres": "", "weight": ""}
"""


@st.cache_resource(show_spinner=False)
def get_gemini_client():
    """Create (and cache) the Gemini client using the API key from st.secrets."""
    try:
        api_key = st.secrets["GEMINI_API_KEY"]
    except Exception:
        return None
    if not api_key:
        return None
    try:
        from google import genai
        return genai.Client(api_key=api_key)
    except Exception as exc:  # pragma: no cover - defensive
        st.error(f"Failed to initialize Gemini client: {exc}")
        return None


def run_ocr(image_bytes: bytes):
    """Send the image to Gemini 2.5 Flash and return a parsed dict, or None on failure."""
    client = get_gemini_client()
    if client is None:
        st.error(
            "⚠️ Gemini API key not found. Please add GEMINI_API_KEY to your "
            "`.streamlit/secrets.toml` file to enable AI tag scanning."
        )
        return None

    try:
        from google.genai import types

        image_part = types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[OCR_PROMPT, image_part],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
        )

        raw_text = (response.text or "").strip()
        if not raw_text:
            st.error("The AI returned an empty response. Please try again with a clearer photo.")
            return None

        # Defensive cleanup in case markdown fences slip through
        if raw_text.startswith("```"):
            raw_text = raw_text.strip("`")
            if raw_text.lower().startswith("json"):
                raw_text = raw_text[4:]
            raw_text = raw_text.strip()

        data = json.loads(raw_text)
        if not isinstance(data, dict):
            st.error("Unexpected AI response format. Please try again.")
            return None
        return data

    except json.JSONDecodeError as exc:
        st.error(f"Failed to parse the AI response as JSON: {exc}")
        return None
    except Exception as exc:
        st.error(f"Gemini OCR request failed: {exc}")
        return None


# ---------------------------------------------------------------------------
# STOCK-IN PAGE
# ---------------------------------------------------------------------------
def init_form_state() -> None:
    defaults = {
        "inp_fabric": "",
        "inp_shade": "",
        "inp_pc_no": "",
        "inp_lot_no": "",
        "inp_metres": "0",
        "inp_weight": "0",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value
    if "last_processed_hash" not in st.session_state:
        st.session_state.last_processed_hash = None
    if "current_image_bytes" not in st.session_state:
        st.session_state.current_image_bytes = None


def clear_form_state() -> None:
    st.session_state.inp_fabric = ""
    st.session_state.inp_shade = ""
    st.session_state.inp_pc_no = ""
    st.session_state.inp_lot_no = ""
    st.session_state.inp_metres = "0"
    st.session_state.inp_weight = "0"
    st.session_state.last_processed_hash = None
    st.session_state.current_image_bytes = None


def stock_in_page() -> None:
    st.header("📸 Stock-In — Capture Roll & Save to Inventory")
    init_form_state()

    conn = get_connection()
    try:
        suppliers = [
            row["supplier_name"]
            for row in conn.execute(
                "SELECT supplier_name FROM suppliers ORDER BY supplier_name"
            ).fetchall()
        ]
    finally:
        conn.close()

    if not suppliers:
        st.warning("No suppliers found. Please add a supplier first in **Manage Suppliers**.")
        return

    st.caption("Capture or upload the fabric roll's identification tag. The AI will read it automatically.")

    input_method = st.radio(
        "Image Source", ["📷 Camera", "🖼️ Gallery Upload"], horizontal=True, key="input_method_radio"
    )

    image_bytes = None
    if input_method == "📷 Camera":
        camera_file = st.camera_input("Capture Roll Tag")
        if camera_file is not None:
            image_bytes = camera_file.getvalue()
    else:
        uploaded_file = st.file_uploader(
            "Upload Roll Tag Image", type=["jpg", "jpeg", "png", "webp"]
        )
        if uploaded_file is not None:
            image_bytes = uploaded_file.getvalue()

    if image_bytes is not None:
        try:
            jpeg_bytes = to_jpeg_bytes(image_bytes)
        except Exception as exc:
            st.error(f"Could not read the captured image: {exc}")
            jpeg_bytes = None

        if jpeg_bytes is not None:
            img_hash = hashlib_md5(jpeg_bytes)
            st.session_state.current_image_bytes = jpeg_bytes

            # Only run OCR once per unique image — this is the core fix for the
            # "widget doesn't update after camera capture" lifecycle issue.
            if img_hash != st.session_state.last_processed_hash:
                with st.spinner("🤖 Reading tag with Gemini AI..."):
                    result = run_ocr(jpeg_bytes)

                st.session_state.last_processed_hash = img_hash

                if result:
                    st.session_state.inp_fabric = str(result.get("fabric") or "")
                    st.session_state.inp_shade = str(result.get("shade") or "")
                    st.session_state.inp_pc_no = str(result.get("pc_no") or "")
                    st.session_state.inp_lot_no = str(result.get("lot_no") or "")
                    st.session_state.inp_metres = clean_numeric_string(result.get("metres"))
                    st.session_state.inp_weight = clean_numeric_string(result.get("weight"))
                    st.success("✅ Tag scanned successfully! Fields auto-filled below.")
                    st.rerun()

            st.image(jpeg_bytes, caption="Captured Roll Image", width=300)

    st.markdown("---")
    st.subheader("Roll Details")

    with st.form("stock_in_form", clear_on_submit=False):
        col1, col2 = st.columns(2)
        with col1:
            supplier = st.selectbox("Supplier", suppliers)
            st.text_input("Fabric Name", key="inp_fabric")
            st.text_input("Color / Shade", key="inp_shade")
        with col2:
            st.text_input("PC No.", key="inp_pc_no")
            st.text_input("Lot No.", key="inp_lot_no")

        col3, col4 = st.columns(2)
        with col3:
            st.text_input("Metres", key="inp_metres")
        with col4:
            st.text_input("Weight (kg)", key="inp_weight")

        submitted = st.form_submit_button(
            "💾 Save Roll to Inventory", use_container_width=True, type="primary"
        )

        if submitted:
            errors = []
            fabric = st.session_state.inp_fabric.strip()
            shade = st.session_state.inp_shade.strip()
            pc_no = st.session_state.inp_pc_no.strip()
            lot_no = st.session_state.inp_lot_no.strip()

            if not fabric:
                errors.append("Fabric Name is required.")

            metres_raw = clean_numeric_string(st.session_state.inp_metres)
            weight_raw = clean_numeric_string(st.session_state.inp_weight)
            metres_val = safe_float(metres_raw, 0.0)
            weight_val = safe_float(weight_raw, 0.0)

            if metres_val < 0 or weight_val < 0:
                errors.append("Metres and Weight cannot be negative.")

            if errors:
                for err in errors:
                    st.error(err)
            else:
                roll_id = generate_roll_id()
                image_path = ""

                if st.session_state.current_image_bytes is not None:
                    try:
                        os.makedirs(IMAGE_DIR, exist_ok=True)
                        image_path = os.path.join(IMAGE_DIR, f"{roll_id}.jpg")
                        with open(image_path, "wb") as f:
                            f.write(st.session_state.current_image_bytes)
                    except Exception as exc:
                        st.warning(f"Could not save roll image to disk: {exc}")
                        image_path = ""

                try:
                    conn = get_connection()
                    conn.execute(
                        """
                        INSERT INTO inventory
                            (roll_id, supplier_name, fabric_name, color_shade, pc_no,
                             lot_no, metres, weight_kg, image_path, status, date_added)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'IN_STOCK', ?)
                        """,
                        (
                            roll_id,
                            supplier,
                            fabric,
                            shade,
                            pc_no,
                            lot_no,
                            metres_val,
                            weight_val,
                            image_path,
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    conn.commit()
                    conn.close()

                    st.success(f"✅ Roll **{roll_id}** saved successfully to inventory!")
                    clear_form_state()
                    st.rerun()

                except sqlite3.IntegrityError as exc:
                    st.error(f"Database integrity error: {exc}")
                except Exception as exc:
                    st.error(f"Unexpected error while saving roll: {exc}")


def hashlib_md5(data: bytes) -> str:
    import hashlib
    return hashlib.md5(data).hexdigest()


# ---------------------------------------------------------------------------
# STOCK-OUT PAGE
# ---------------------------------------------------------------------------
def stock_out_page() -> None:
    st.header("➖ Stock-Out — Dispatch Roll")

    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM inventory WHERE status = 'IN_STOCK' ORDER BY date_added DESC"
        ).fetchall()
    finally:
        conn.close()

    if not rows:
        st.info("No rolls currently in stock.")
        return

    options = {
        f"{r['roll_id']} — {r['fabric_name']} ({r['color_shade']}) — {r['metres']}m": r["roll_id"]
        for r in rows
    }
    selected_label = st.selectbox("Select Roll to Dispatch", list(options.keys()))
    selected_roll_id = options[selected_label]
    selected_row = next(r for r in rows if r["roll_id"] == selected_roll_id)

    col1, col2 = st.columns([1, 2])
    with col1:
        if selected_row["image_path"] and os.path.exists(selected_row["image_path"]):
            st.image(selected_row["image_path"], width=250)
        else:
            st.info("No image available")
    with col2:
        st.markdown(f"<span class='roll-id-badge'>{selected_row['roll_id']}</span>", unsafe_allow_html=True)
        st.write("")
        st.write(f"**Supplier:** {selected_row['supplier_name']}")
        st.write(f"**Fabric:** {selected_row['fabric_name']}")
        st.write(f"**Shade:** {selected_row['color_shade']}")
        st.write(f"**PC No.:** {selected_row['pc_no']}")
        st.write(f"**Lot No.:** {selected_row['lot_no']}")
        st.write(f"**Metres:** {selected_row['metres']}")
        st.write(f"**Weight (kg):** {selected_row['weight_kg']}")
        st.write(f"**Date Added:** {selected_row['date_added']}")

    st.markdown("---")
    confirm = st.checkbox(f"I confirm dispatch of roll {selected_row['roll_id']}")
    if st.button("🚚 Confirm Dispatch", type="primary", use_container_width=True, disabled=not confirm):
        try:
            conn = get_connection()
            conn.execute(
                "UPDATE inventory SET status = 'DISPATCHED' WHERE roll_id = ?",
                (selected_roll_id,),
            )
            conn.commit()
            conn.close()
            st.success(f"✅ Roll {selected_roll_id} has been dispatched.")
            st.rerun()
        except Exception as exc:
            st.error(f"Error dispatching roll: {exc}")


# ---------------------------------------------------------------------------
# MANAGE SUPPLIERS PAGE
# ---------------------------------------------------------------------------
def manage_suppliers_page() -> None:
    st.header("🏢 Manage Suppliers")

    with st.form("add_supplier_form", clear_on_submit=True):
        name = st.text_input("Supplier Name")
        desc = st.text_area("Description (optional)")
        submitted = st.form_submit_button("➕ Add Supplier", type="primary")

        if submitted:
            clean_name = name.strip()
            if not clean_name:
                st.error("Supplier name is required.")
            else:
                try:
                    conn = get_connection()
                    conn.execute(
                        "INSERT INTO suppliers (supplier_name, description) VALUES (?, ?)",
                        (clean_name, desc.strip()),
                    )
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Supplier '{clean_name}' added.")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error(f"Supplier '{clean_name}' already exists.")
                except Exception as exc:
                    st.error(f"Error adding supplier: {exc}")

    st.markdown("---")
    st.subheader("Existing Suppliers")

    conn = get_connection()
    try:
        suppliers = conn.execute("SELECT * FROM suppliers ORDER BY supplier_name").fetchall()
    finally:
        conn.close()

    if suppliers:
        data = [
            {
                "ID": s["id"],
                "Supplier Name": s["supplier_name"],
                "Description": s["description"] or "",
            }
            for s in suppliers
        ]
        st.dataframe(data, use_container_width=True, hide_index=True)
    else:
        st.info("No suppliers found.")


# ---------------------------------------------------------------------------
# INVENTORY SHEET & EXCEL EXPORT PAGE
# ---------------------------------------------------------------------------
def build_excel_export(rows) -> bytes:
    """Build an .xlsx file with embedded roll thumbnail images using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "Image", "Roll ID", "Supplier", "Fabric", "Shade", "PC No.",
        "Lot No.", "Metres", "Weight (kg)", "Status", "Date Added",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [14, 20, 16, 18, 16, 12, 12, 10, 12, 12, 18]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    row_num = 2
    for r in rows:
        ws.row_dimensions[row_num].height = 62

        ws.cell(row=row_num, column=2, value=r["roll_id"])
        ws.cell(row=row_num, column=3, value=r["supplier_name"])
        ws.cell(row=row_num, column=4, value=r["fabric_name"])
        ws.cell(row=row_num, column=5, value=r["color_shade"])
        ws.cell(row=row_num, column=6, value=r["pc_no"])
        ws.cell(row=row_num, column=7, value=r["lot_no"])
        ws.cell(row=row_num, column=8, value=r["metres"])
        ws.cell(row=row_num, column=9, value=r["weight_kg"])
        ws.cell(row=row_num, column=10, value=r["status"])
        ws.cell(row=row_num, column=11, value=r["date_added"])

        image_path = r["image_path"]
        if image_path and os.path.exists(image_path):
            try:
                thumb = Image.open(image_path)
                thumb.thumbnail((100, 100))
                if thumb.mode != "RGB":
                    thumb = thumb.convert("RGB")
                xl_img = XLImage(thumb)
                xl_img.width = 82
                xl_img.height = 82
                ws.add_image(xl_img, f"A{row_num}")
            except Exception:
                ws.cell(row=row_num, column=1, value="N/A")
        else:
            ws.cell(row=row_num, column=1, value="No Image")

        row_num += 1

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def inventory_sheet_page() -> None:
    st.header("📊 Inventory Sheet & Excel Export")

    conn = get_connection()
    try:
        all_rows = conn.execute("SELECT * FROM inventory ORDER BY date_added DESC").fetchall()
        suppliers = ["All"] + [
            row["supplier_name"]
            for row in conn.execute(
                "SELECT supplier_name FROM suppliers ORDER BY supplier_name"
            ).fetchall()
        ]
    finally:
        conn.close()

    if not all_rows:
        st.info("No inventory records found yet. Start by scanning a roll in **Stock-In**.")
        return

    col1, col2, col3 = st.columns(3)
    with col1:
        status_filter = st.selectbox("Filter by Status", ["All", "IN_STOCK", "DISPATCHED"])
    with col2:
        supplier_filter = st.selectbox("Filter by Supplier", suppliers)
    with col3:
        search_term = st.text_input("Search (Fabric / Lot / PC No.)", "")

    filtered = []
    for r in all_rows:
        if status_filter != "All" and r["status"] != status_filter:
            continue
        if supplier_filter != "All" and r["supplier_name"] != supplier_filter:
            continue
        if search_term.strip():
            term = search_term.strip().lower()
            haystack = " ".join(
                str(x or "")
                for x in [r["fabric_name"], r["lot_no"], r["pc_no"], r["color_shade"]]
            ).lower()
            if term not in haystack:
                continue
        filtered.append(r)

    st.write(f"**{len(filtered)}** record(s) found.")

    table_data = [
        {
            "Roll ID": r["roll_id"],
            "Supplier": r["supplier_name"],
            "Fabric": r["fabric_name"],
            "Shade": r["color_shade"],
            "PC No.": r["pc_no"],
            "Lot No.": r["lot_no"],
            "Metres": r["metres"],
            "Weight (kg)": r["weight_kg"],
            "Status": r["status"],
            "Date Added": r["date_added"],
        }
        for r in filtered
    ]
    st.dataframe(table_data, use_container_width=True, hide_index=True)

    st.markdown("---")
    if st.button("📥 Generate Excel Export (with Thumbnails)", type="primary"):
        if not filtered:
            st.warning("No records to export with the current filters.")
        else:
            with st.spinner("Building Excel file..."):
                try:
                    excel_bytes = build_excel_export(filtered)
                except Exception as exc:
                    st.error(f"Failed to build Excel export: {exc}")
                    excel_bytes = None

            if excel_bytes:
                st.download_button(
                    label="⬇️ Download Excel File",
                    data=excel_bytes,
                    file_name=f"inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


# ---------------------------------------------------------------------------
# MAIN / NAVIGATION
# ---------------------------------------------------------------------------
def main() -> None:
    init_db()

    st.sidebar.title("🧵 Fabric Warehouse")
    st.sidebar.markdown("---")
    page = st.sidebar.radio(
        "Navigation",
        [
            "📸 Stock-In",
            "➖ Stock-Out",
            "🏢 Manage Suppliers",
            "📊 Inventory Sheet & Excel",
        ],
    )
    st.sidebar.markdown("---")

    try:
        conn = get_connection()
        in_stock_count = conn.execute(
            "SELECT COUNT(*) FROM inventory WHERE status = 'IN_STOCK'"
        ).fetchone()[0]
        dispatched_count = conn.execute(
            "SELECT COUNT(*) FROM inventory WHERE status = 'DISPATCHED'"
        ).fetchone()[0]
        conn.close()
        st.sidebar.metric("Rolls In Stock", in_stock_count)
        st.sidebar.metric("Rolls Dispatched", dispatched_count)
    except Exception:
        pass

    st.sidebar.markdown("---")
    st.sidebar.caption("Powered by Gemini 2.5 Flash AI OCR")

    st.title("Fabric Warehouse Inventory Tracking")

    if page == "📸 Stock-In":
        stock_in_page()
    elif page == "➖ Stock-Out":
        stock_out_page()
    elif page == "🏢 Manage Suppliers":
        manage_suppliers_page()
    elif page == "📊 Inventory Sheet & Excel":
        inventory_sheet_page()


if __name__ == "__main__":
    main()
